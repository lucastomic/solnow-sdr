#!/usr/bin/env python3
"""
SolNow — Enriched Pipeline Orchestrator

Ties together: Google Maps search → Web enrichment with AI → GMV estimation → ICP scoring.
Outputs a fully enriched Excel/CSV for SDR prospecting.
"""

import asyncio
import logging
import os
import sys
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from enrichment import enrich_all
from gmv import score_and_estimate
from prospect import DEFAULT_ZONES, QUERIES, get_first_review_date, run_search

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ── Output columns (enriched) ────────────────────────────────────────────────

ENRICHED_COLUMNS = [
    "Nombre",
    "Dirección",
    "Zona Costera",
    "Teléfono",
    "Web",
    "Google Rating",
    "Reseñas Total",
    "Fecha 1ª Reseña",
    "Reseñas/Año",
    "Categoría",
    "Tipo Precio",
    "Precio Medio (€)",
    "Nº Activos",
    "Capacidad/Salida",
    "Días Temporada",
    "GMV Estimado Anual (€)",
    "Comisión Anual (€)",
    "Comisión Mensual (€)",
    "Score ICP",
    "Es ICP",
]


# ── Flatten Google Places results to operator dicts ──────────────────────────

def _flatten_results(zone_results: dict) -> list[dict]:
    """Convert zone_results {zone: {pid: place}} to a flat list of operator dicts."""
    operators = []
    for zona, places in zone_results.items():
        for pid, place in places.items():
            first_review = get_first_review_date(place)

            # Calculate years active from first review
            años_activo = 1.0
            if first_review:
                try:
                    first_dt = datetime.fromisoformat(first_review)
                    now = datetime.now(timezone.utc)
                    if first_dt.tzinfo is None:
                        first_dt = first_dt.replace(tzinfo=timezone.utc)
                    diff = (now - first_dt).days / 365.25
                    años_activo = max(diff, 0.5)
                except (ValueError, TypeError):
                    pass

            reviews_total = place.get("userRatingCount") or 0
            reviews_por_año = round(reviews_total / max(años_activo, 1), 1)

            op = {
                "place_id": pid,
                "nombre": place.get("displayName", {}).get("text", ""),
                "direccion": place.get("formattedAddress", ""),
                "zona": zona,
                "telefono": (
                    place.get("nationalPhoneNumber")
                    or place.get("internationalPhoneNumber", "")
                ),
                "web": place.get("websiteUri", ""),
                "google_rating": place.get("rating", ""),
                "google_reviews_total": reviews_total,
                "fecha_primera_review": first_review,
                "años_activo": round(años_activo, 1),
                "reviews_por_año": reviews_por_año,
                "price_level": place.get("priceLevel", ""),
            }
            operators.append(op)
    return operators


def _operator_to_row(op: dict) -> list:
    """Convert an enriched operator dict to an Excel row."""
    return [
        op.get("nombre", ""),
        op.get("direccion", ""),
        op.get("zona_costera", ""),
        op.get("telefono", ""),
        op.get("web", ""),
        op.get("google_rating", ""),
        op.get("google_reviews_total", ""),
        op.get("fecha_primera_review", ""),
        op.get("reviews_por_año", ""),
        op.get("categoria_principal", ""),
        op.get("tipo_precio", ""),
        op.get("precio_medio", ""),
        op.get("num_activos", ""),
        op.get("capacidad_maxima_por_salida", ""),
        op.get("dias_temporada", ""),
        op.get("gmv_estimado_anual", ""),
        op.get("comision_solnow_anual", ""),
        op.get("comision_solnow_mensual", ""),
        op.get("score_icp", ""),
        "Sí" if op.get("es_icp") else "No",
    ]


# ── Pipeline ─────────────────────────────────────────────────────────────────

async def run_pipeline(
    google_api_key: str,
    anthropic_api_key: str | None = None,
    zones: list[str] | None = None,
    queries: list[str] | None = None,
    on_place=None,
    on_enrich_progress=None,
) -> list[dict]:
    """Run the full enrichment pipeline.

    Args:
        google_api_key: Google Places API key.
        anthropic_api_key: Anthropic API key (optional — skip enrichment if None).
        zones: List of zones to search.
        queries: Custom query templates (uses defaults if None).
        on_place: Callback(pid, place) for each place found during search.
        on_enrich_progress: Callback(current, total, name) for enrichment progress.

    Returns:
        List of enriched operator dicts sorted by score_icp DESC.
    """
    zones = zones or DEFAULT_ZONES
    log.info("Starting pipeline: %d zones, enrichment=%s", len(zones), bool(anthropic_api_key))

    # Phase 1: Google Maps search
    log.info("Phase 1: Google Maps search...")
    zone_results = run_search(google_api_key, zones, on_place=on_place, queries=queries)
    operators = _flatten_results(zone_results)
    log.info("Found %d operators total", len(operators))

    # Phase 2: Web enrichment with AI (if API key provided)
    if anthropic_api_key and operators:
        log.info("Phase 2: Web enrichment with Claude Haiku...")
        await enrich_all(operators, anthropic_api_key, on_progress=on_enrich_progress)
        log.info("Enrichment complete")

        # Filter out operators classified as non-aquatic by name inference
        before = len(operators)
        operators = [
            op for op in operators
            if op.get("categoria_principal") != "no_acuatico"
        ]
        removed = before - len(operators)
        if removed:
            log.info("Removed %d non-aquatic operators (no_acuatico)", removed)
    else:
        log.info("Phase 2: Skipped (no Anthropic API key)")

    # Phase 3: GMV estimation + ICP scoring
    log.info("Phase 3: GMV estimation + ICP scoring...")
    for op in operators:
        score_and_estimate(op)

    # Remove marketplaces from output (they contaminate the pipeline)
    before = len(operators)
    operators = [op for op in operators if not op.get("es_marketplace")]
    marketplace_count = before - len(operators)
    if marketplace_count:
        log.info("Removed %d marketplace/intermediary operators", marketplace_count)

    # Sort: best prospects first
    operators.sort(key=lambda o: (-o.get("score_icp", 0), -o.get("comision_solnow_mensual", 0)))

    log.info("Pipeline complete. %d operators, %d ICP",
             len(operators), sum(1 for o in operators if o.get("es_icp")))

    return operators


# ── Export ────────────────────────────────────────────────────────────────────

def _autofit(ws):
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def export_enriched(operators: list[dict], filename: str):
    """Export enriched operators to an Excel file."""
    wb = Workbook()
    wb.remove(wb.active)

    # "All" sheet — sorted by score
    ws_all = wb.create_sheet(title="All", index=0)
    ws_all.append(ENRICHED_COLUMNS)
    for op in operators:
        ws_all.append(_operator_to_row(op))
    _autofit(ws_all)

    # "ICP Only" sheet
    icp_ops = [o for o in operators if o.get("es_icp")]
    if icp_ops:
        ws_icp = wb.create_sheet(title="ICP")
        ws_icp.append(ENRICHED_COLUMNS)
        for op in icp_ops:
            ws_icp.append(_operator_to_row(op))
        _autofit(ws_icp)

    # Per-zone sheets
    zones_seen = {}
    for op in operators:
        z = op.get("zona_costera") or op.get("zona", "Otros")
        zones_seen.setdefault(z, []).append(op)

    for zona, ops in zones_seen.items():
        title = zona[:31].replace("/", "-")
        ws = wb.create_sheet(title=title)
        ws.append(ENRICHED_COLUMNS)
        for op in ops:
            ws.append(_operator_to_row(op))
        _autofit(ws)

    wb.save(filename)
    log.info("Exported enriched data to %s (%d operators, %d ICP)",
             filename, len(operators), len(icp_ops))


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    google_key = os.environ.get("GOOGLE_PLACES_API_KEY")
    anthropic_key = os.environ.get("ANTHROPIC_API_KEY")

    if not google_key:
        print("Error: set GOOGLE_PLACES_API_KEY environment variable")
        sys.exit(1)

    zones = sys.argv[1:] if len(sys.argv) > 1 else DEFAULT_ZONES

    operators = asyncio.run(run_pipeline(
        google_api_key=google_key,
        anthropic_api_key=anthropic_key,
        zones=zones,
    ))

    total = len(operators)
    icp_count = sum(1 for o in operators if o.get("es_icp"))
    filename = "prospects_enriched.xlsx"
    export_enriched(operators, filename)

    print("\n--- Summary ---")
    print(f"  Total operators: {total}")
    print(f"  ICP prospects:   {icp_count}")
    if anthropic_key:
        print(f"  Enriched:        {sum(1 for o in operators if o.get('categoria_principal'))}")
    else:
        print("  Enrichment:      Skipped (set ANTHROPIC_API_KEY to enable)")
    print(f"  Output:          {filename}")


if __name__ == "__main__":
    main()
