#!/usr/bin/env python3
"""
SolNow Prospecting Script
Searches for jet ski / personal watercraft rental companies in Spain using Google Places API (New)
and exports results to Excel for SDR outreach.
"""

import logging
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

QUERIES = [
    # Jet ski
    "alquiler motos de agua {zona}",
    "jet ski alquiler {zona}",
    "motos acuáticas {zona}",
    # Kayak / paddle
    "alquiler kayak {zona}",
    "kayak guiado {zona}",
    "paddle surf alquiler {zona}",
    "SUP alquiler {zona}",
    # Boat excursions
    "excursiones en barco {zona}",
    "paseos en barco {zona}",
    "alquiler barco sin licencia {zona}",
    "barco compartido excursion {zona}",
    # Charter
    "charter nautico {zona}",
    "alquiler velero {zona}",
    "alquiler catamaran {zona}",
    # Generic
    "actividades acuáticas {zona}",
    "deportes acuáticos {zona}",
    "experiencias nauticas {zona}",
    "water sports {zona}",
]

DEFAULT_ZONES = [
    # Costa del Sol
    "Marbella", "Málaga", "Torremolinos", "Nerja", "Estepona",
    # Costa Blanca
    "Alicante", "Benidorm", "Calpe", "Denia", "Jávea",
    # Comunidad Valenciana
    "Valencia", "Gandía", "Cullera",
    # Baleares
    "Ibiza", "Formentera", "Palma de Mallorca", "Menorca",
    # Costa Brava / Maresme
    "Barcelona", "Sitges", "Lloret de Mar", "Roses",
    # Canarias
    "Tenerife", "Gran Canaria", "Lanzarote", "Fuerteventura",
    # Costa de la Luz
    "Tarifa", "Cádiz", "Conil",
    # Murcia
    "Murcia", "Cartagena", "La Manga",
]

EXCLUDE_KEYWORDS = [
    "escuela",
    "club náutico",
    "club nautico",
    "ferry",
    "ferri",
    "academia",
    "federación",
    "federacion",
    "puerto deportivo",
    "capitanía",
    "capitania",
    "concesionario",
    "taller",
    "reparación",
    "reparacion",
    "tienda de recambios",
    # Retail / commercial
    "decathlon",
    "centro comercial",
    "supermercado",
    "inmobiliaria",
    "real estate",
    "agencia de viajes",
    # Non-water tourism
    "free tour",
    "tour a pie",
    "mirador",
    "escape room",
    "paintball",
    "karting",
    "bowling",
    "cine ",
    "teatro",
    # Gyms / sports clubs
    "gimnasio",
    "fitness",
    "activaclub",
    "sport club",
    # Hospitality (not operators)
    "spa ",
    "hotel ",
    "restaurante",
    "bar ",
    "cafeteria",
    # Parks / attractions
    "parque de ",
    "parque municipal",
    "ilusiona",
    # Misc
    "tinglado",
]

# Google Maps types that indicate a valid water sports operator
VALID_GOOGLE_TYPES = [
    "water_sports",
    "boat_rental",
    "boat_tour",
    "scuba_diving",
    "kayak_rental",
    "jet_ski_rental",
    "sailing",
    "tour_operator",
    "tourist_attraction",
    "travel_agency",
    "amusement_center",
]

# Keywords in Google type display names (localized) that are valid
VALID_CATEGORY_KEYWORDS = [
    "náutic", "nautic", "barco", "embarcación", "embarcacion", "charter",
    "alquiler de barcos", "water sport", "deportes acuáticos", "deportes acuaticos",
    "actividades acuáticas", "actividades acuaticas", "kayak", "paddle", "jet ski",
    "motos de agua", "buceo", "vela", "sailing", "boat",
    "excursión marítima", "excursion maritima", "paseo en barco",
    "diving", "snorkel", "surf", "canoa", "piragua",
]

COLUMNS = [
    "Nombre", "Dirección", "Teléfono", "Web", "Rating", "Reseñas",
    "Zona", "Nivel Precio", "Fecha 1ª Reseña", "Estado",
]

SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"

SEARCH_FIELDS = (
    "nextPageToken,"
    "places.id,"
    "places.displayName,"
    "places.formattedAddress,"
    "places.nationalPhoneNumber,"
    "places.internationalPhoneNumber,"
    "places.websiteUri,"
    "places.rating,"
    "places.userRatingCount,"
    "places.currentOpeningHours,"
    "places.priceLevel,"
    "places.reviews,"
    "places.types,"
    "places.primaryTypeDisplayName"
)


def is_relevant(place):
    """Filter out irrelevant results using two-layer validation.

    Layer 1: Exclude by name keywords (Decathlon, gyms, malls, etc.)
    Layer 2: Validate by Google Maps type/category (water sports related)
    If type data is missing, include the result (benefit of the doubt).
    """
    if not place.get("websiteUri") and not place.get("nationalPhoneNumber"):
        return False

    name_lower = place.get("displayName", {}).get("text", "").lower()

    # Layer 1: Exclude obvious non-operators by name
    if any(kw in name_lower for kw in EXCLUDE_KEYWORDS):
        return False

    # Layer 2: Validate by Google Maps category
    # Get all type signals from the place
    types = place.get("types") or []
    primary_display = (place.get("primaryTypeDisplayName") or {}).get("text", "").lower()

    # If we have type data, check if any type matches valid water sports categories
    if types or primary_display:
        # Check structured types
        has_valid_type = any(t in VALID_GOOGLE_TYPES for t in types)

        # Check display name keywords (localized category names)
        has_valid_keyword = any(kw in primary_display for kw in VALID_CATEGORY_KEYWORDS)

        # Also check name for water-sports signals as a safety net
        name_has_water_signal = any(kw in name_lower for kw in VALID_CATEGORY_KEYWORDS)

        if has_valid_type or has_valid_keyword or name_has_water_signal:
            return True

        # Has type data but none match → likely not a water sports operator
        log.debug("Excluded (no valid category): %s [types=%s, primary=%s]",
                   name_lower, types, primary_display)
        return False

    # No type data at all → benefit of the doubt, include it
    return True


def text_search(api_key, query, page_token=None):
    """Call Places API (New) Text Search."""
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": SEARCH_FIELDS,
    }
    body = {"textQuery": query, "languageCode": "es"}
    if page_token:
        body["pageToken"] = page_token

    resp = requests.post(SEARCH_URL, json=body, headers=headers)
    if not resp.ok:
        log.error("API error %s: %s", resp.status_code, resp.text)
        resp.raise_for_status()
    return resp.json()


def search_single_query(api_key, query, zona, on_place=None):
    """Execute a single query with pagination, returning {place_id: place} dict."""
    results = {}
    log.info("Searching: %s", query)

    data = text_search(api_key, query)
    pages_fetched = 1

    while True:
        for place in data.get("places", []):
            pid = place.get("id")
            if pid in results:
                continue
            place["_zona"] = zona
            if is_relevant(place):
                results[pid] = place
                name = place.get("displayName", {}).get("text", "?")
                log.info("  + %s", name)
                if on_place:
                    on_place(pid, place)

        token = data.get("nextPageToken")
        if not token or pages_fetched >= 5:
            break

        time.sleep(2)  # Required by Google for nextPageToken
        data = text_search(api_key, query, page_token=token)
        pages_fetched += 1

    return results


def search_zone(api_key, zona, on_place=None, queries=None):
    """Search a zone with multiple queries in parallel, returning deduplicated results."""
    results = {}
    templates = queries if queries else QUERIES
    expanded = [tpl.format(zona=zona) if "{zona}" in tpl else f"{tpl} {zona}" for tpl in templates]

    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(search_single_query, api_key, q, zona, on_place): q
            for q in expanded
        }
        for future in as_completed(futures):
            try:
                partial = future.result()
                for pid, place in partial.items():
                    if pid not in results:
                        results[pid] = place
            except Exception:
                query = futures[future]
                log.exception("Query failed: %s", query)

    return results


def run_search(api_key, zones, on_place=None, queries=None):
    """Run multi-zone search. Returns deduplicated zone_results dict.

    Args:
        api_key: Google Places API key
        zones: List of zone names to search
        on_place: Optional callback(pid, place) called for each new unique place found
    """
    raw_results = {}
    with ThreadPoolExecutor(max_workers=len(zones)) as executor:
        futures = {
            executor.submit(search_zone, api_key, zona, on_place, queries): zona
            for zona in zones
        }
        for future in as_completed(futures):
            zona = futures[future]
            log.info("=== Zone: %s ===", zona)
            try:
                raw_results[zona] = future.result()
            except Exception:
                log.exception("Zone failed: %s", zona)
                raw_results[zona] = {}

    # Cross-zone deduplication (preserves order of zones as given)
    zone_results = {}
    global_seen = set()
    for zona in zones:
        results = raw_results.get(zona, {})
        unique = {}
        for pid, place in results.items():
            if pid not in global_seen:
                global_seen.add(pid)
                unique[pid] = place
        zone_results[zona] = unique
        log.info("Zone %s: %d leads", zona, len(unique))

    return zone_results


def get_first_review_date(place):
    """Extract the date of the earliest review from the reviews list."""
    reviews = place.get("reviews") or []
    dates = []
    for r in reviews:
        ts = r.get("publishTime") or r.get("relativePublishTimeDescription", "")
        if ts and "T" in str(ts):  # ISO-like timestamp
            dates.append(str(ts)[:10])
    return min(dates) if dates else ""


def place_to_row(place):
    return [
        place.get("displayName", {}).get("text", ""),
        place.get("formattedAddress", ""),
        place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber", ""),
        place.get("websiteUri", ""),
        place.get("rating", ""),
        place.get("userRatingCount", ""),
        place.get("_zona", ""),
        place.get("priceLevel", ""),
        get_first_review_date(place),
        "",  # Estado — empty for SDR
    ]


def autofit_columns(ws):
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        max_len = 0
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def export_excel(zone_results, filename):
    wb = Workbook()
    wb.remove(wb.active)

    all_rows = []

    for zona, places in zone_results.items():
        ws = wb.create_sheet(title=zona[:31])
        ws.append(COLUMNS)
        for place in places.values():
            row = place_to_row(place)
            ws.append(row)
            all_rows.append(row)
        autofit_columns(ws)

    ws_all = wb.create_sheet(title="All", index=0)
    ws_all.append(COLUMNS)
    for row in all_rows:
        ws_all.append(row)
    autofit_columns(ws_all)

    wb.save(filename)
    log.info("Exported to %s", filename)


def main():
    api_key = os.environ.get("GOOGLE_PLACES_API_KEY")
    if not api_key:
        print("Error: set GOOGLE_PLACES_API_KEY environment variable")
        sys.exit(1)

    zones = sys.argv[1:] if len(sys.argv) > 1 else DEFAULT_ZONES

    zone_results = run_search(api_key, zones)

    total = sum(len(v) for v in zone_results.values())
    filename = "prospects_solnow.xlsx"
    export_excel(zone_results, filename)

    print("\n--- Summary ---")
    for zona, places in zone_results.items():
        print(f"  {zona}: {len(places)} leads")
    print(f"  Total: {total} leads")
    print(f"  Output: {filename}")


if __name__ == "__main__":
    main()
